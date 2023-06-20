
import datetime
import openpyxl
from ortools.sat.python import cp_model

'''
################################
||         PARAMETER          ||
################################
'''
EXCEL_INPUT_PATH = "inputs/inputSE.xlsx"
MAX_EXCECUTE_TIME = 120.0
OBJECTIVE_STATUS = [1,0,0,0,0,0]

class IntructorAssigningSolver():
    def __init__(self,max_time_per_obj,excel_input_path,obj_status):
        self.MAX_CP_TIME = max_time_per_obj
        self.EXCEL_INPUT_PATH = excel_input_path
        self.OBJ_STATUS = obj_status
        '''
        ################################
        ||           SHEET            ||
        ################################
        '''
        data_wb = openpyxl.load_workbook(self.EXCEL_INPUT_PATH)
        self.info_sheet = data_wb.worksheets[0]
        self.task_sheet = data_wb.worksheets[1]
        self.slot_conflict_sheet = data_wb.worksheets[2]
        self.slot_compatibility_sheet = data_wb.worksheets[3]
        self.instructor_ability_sheet = data_wb.worksheets[4]
        self.instructor_slot_sheet = data_wb.worksheets[5]
        self.instructor_quota_sheet = data_wb.worksheets[6]
        self.instructor_preassign_sheet = data_wb.worksheets[7]
        self.area_distance_sheet = data_wb.worksheets[8]
        self.area_slot_weight_sheet = data_wb.worksheets[9]

    def create_variable(self):
        '''
        ################################
        ||          VARIABLE          ||
        ################################
        '''
        self.num_task = self.info_sheet.cell(1,2).value
        self.num_instructor = self.info_sheet.cell(2,2).value
        self.num_slot = self.info_sheet.cell(3,2).value
        self.num_subject = self.info_sheet.cell(4,2).value
        self.num_area = self.info_sheet.cell(5,2).value
        self.backup_quota = self.info_sheet.cell(6,2).value

        self.all_subject = range(self.num_subject)
        self.all_task = range(self.num_task)
        self.all_slot = range(self.num_slot)
        self.all_instructor = range(self.num_instructor)
        self.all_area = range(self.num_area)

        self.slot_conflict = [[0] * self.num_slot for i in self.all_slot]
        for i in self.all_slot:
            for j in self.all_slot:
                if self.slot_conflict_sheet.cell(i+2,j+2).value:
                    self.slot_conflict[i][j]=1

        self.slot_compatibility=[[0] * self.num_slot for i in self.all_slot]
        for i in self.all_slot:
            for j in self.all_slot:
                    self.slot_compatibility[i][j]=int(self.slot_compatibility_sheet.cell(i+2,j+2).value)

        self.instructor_ability = [[0] * self.num_subject for i in self.all_instructor]
        self.instructor_subject_preference = [[0] * self.num_subject for i in self.all_instructor]
        for i in self.all_instructor:
            for j in self.all_subject:
                if self.instructor_ability_sheet.cell(i+2,j+2).value:
                    self.instructor_ability[i][j]=1
                self.instructor_subject_preference[i][j]=self.instructor_ability_sheet.cell(i+2,j+2).value
        
        self.instructor_slot = [[0] * self.num_slot for i in self.all_instructor]
        self.instructor_slot_preference= [[0] * self.num_slot for i in self.all_instructor]
        for i in self.all_instructor:
            for j in self.all_slot:
                if self.instructor_slot_sheet.cell(i+2,j+2).value:
                    self.instructor_slot[i][j]=1
                self.instructor_slot_preference[i][j]=self.instructor_slot_sheet.cell(i+2,j+2).value

        self.task_subject_mapping = []
        for i in self.all_task:
            for j in self.all_subject:
                if self.task_sheet.cell(i+2,2).value == self.instructor_ability_sheet.cell(1,j+2).value:
                    self.task_subject_mapping.append(j)

        self.task_slot_mapping = []
        for i in self.all_task:
            for j in self.all_slot:
                if self.task_sheet.cell(i+2,4).value == self.instructor_slot_sheet.cell(1,j+2).value:
                    self.task_slot_mapping.append(j)

        self.instructor_quota=[]
        for i in self.all_instructor:
            self.instructor_quota.append(int ( self.instructor_quota_sheet.cell(i+2,2).value) )
        self.instructor_quota.append(self.backup_quota)

        self.instructor_preassign = [[0] * self.num_task for i in self.all_instructor]
        for i in self.all_instructor:
            for n in self.all_task:
                if self.instructor_preassign_sheet.cell(i+2,n+2).value:
                    self.instructor_preassign[i][self.instructor_preassign_sheet.cell(i+2,n+2).value-1]=1

        self.area_distance = [[0]* self.num_area for i in self.all_area]
        for i in self.all_area:
            for j in self.all_area:
                self.area_distance[i][j] = self.area_distance_sheet.cell(i+2,j+2).value
        
        self.area_slot_weight = [[0] * self.num_slot for i in self.all_slot]
        for i in self.all_slot:
            for j in self.all_slot:
                self.area_slot_weight[i][j]=int(self.area_slot_weight_sheet.cell(i+2,j+2).value)

        self.task_area_mapping = []
        for n in self.all_task:
            dept=self.task_sheet.cell(n+2,7).value[0:2]
            if(dept=="AL"): self.task_area_mapping.append(0)
            if(dept=="BE"): self.task_area_mapping.append(1)
            if(dept=="DE"): self.task_area_mapping.append(2)
        
        self.slot_task_listing = [[] for s in self.all_slot]
        for n in self.all_task:
            self.slot_task_listing[self.task_slot_mapping[n]].append(n)

        self.slot_task_conflicting = [[] for s in self.all_slot]
        for s in self.all_slot:
            for n in self.all_task:
                if(self.slot_conflict[self.task_slot_mapping[n]][s]):
                    self.slot_task_conflicting[s].append(n)
        
    def create_model(self):
        '''
        ################################
        ||           Model            ||
        ################################
        '''
        # Creates the model.
        self.model = cp_model.CpModel()

        # Decision variable
        self.assign = {}
        for n in self.all_task:
            for i in range(self.num_instructor+1):
                self.assign[(n,i)] = self.model.NewBoolVar('shift_n%ii%i' % (n,i))
        
        '''
        ################################
        ||         CONSTRAINT         ||
        ################################
        '''
        #C-00 CONSTRAINT EACH TASK ASSIGN TO ONLY ONE 
        for n in self.all_task:
            self.model.AddExactlyOne(self.assign[(n,i)] for i in range(self.num_instructor+1))

        #C-00 CONSTRAINT INSTRUCTOR QUOTA MUST IN RANGE
        for i in range(self.num_instructor+1):
            task_assigned = []
            for n in self.all_task:
                task_assigned.append( self.assign[(n,i)] )
            self.model.Add( sum( task_assigned ) <= self.instructor_quota[i])
        
        #C-01 CONSTRAINT NOT ASSIGN 2 CONFLICT SLOT
        for i in self.all_instructor:
            for s in self.all_slot:    
                total_task_in_that_slot=[]
                total_task_conflict_with_that_slot=[]
                for n in self.slot_task_listing[s]:
                    total_task_in_that_slot.append(self.assign[(n,i)])
                for n in self.slot_task_conflicting[s]:
                    total_task_conflict_with_that_slot.append(self.assign[(n,i)])
                tmp = self.model.NewBoolVar("")
                self.model.Add(sum(total_task_in_that_slot)>0).OnlyEnforceIf(tmp)
                self.model.Add(sum(total_task_in_that_slot)==0).OnlyEnforceIf(tmp.Not())
                self.model.Add(sum(total_task_conflict_with_that_slot)==1).OnlyEnforceIf(tmp)
        
        #C-02 CONSTRAINT  PREASSIGN MUST BE SATISFY  
        for i in self.all_instructor:
            for s in self.all_slot:
                n = self.instructor_preassign_sheet.cell(i+2,s+2).value
                if n is not None:
                    self.model.Add(self.assign[(n-1,i)] == 1)

        #C-03 CONSTRAINT INSTRUCTOR MUST HAVE ABILITY FOR THAT SUBJECT
        for n in self.all_task:
            for i in self.all_instructor:
                self.model.Add( (self.instructor_ability[i][self.task_subject_mapping[n]]- self.assign[(n,i)]) > -1)

        #C-04 CONSTRAINT INSTRUCTOR MUST BE ABLE TO TEACH IN THAT SLOT
        for n in self.all_task:
            for i in self.all_instructor:
                self.model.Add( (self.instructor_slot[i][self.task_slot_mapping[n]]- self.assign[(n,i)]) > -1)

    '''
    ################################
    ||         OBJECTIVE          ||
    ################################
    '''
    # O-00
    def obj_backup_instructor(self):
        need_backup_instructor_task = []
        for n in self.all_task:
            need_backup_instructor_task.append(self.assign[(n,self.num_instructor)])
        return sum(need_backup_instructor_task)
    
    # O-01
    def obj_slot_compatibility(self):
        objective_slot_compatibility=[]
        for n1 in range(self.num_task-1):
            for n2 in range(n1+1,self.num_task):
                objective_slot_compatibility.append(self.slot_compatibility[self.task_slot_mapping[n1]][self.task_slot_mapping[n2]]*self.boolean_product[(n1,n2)])
        return sum(objective_slot_compatibility)
    
    # O-02
    def obj_subject_diversity(self):
        subject_diversity=[]
        for i in self.all_instructor:
            tmp=[]
            for s in self.all_subject:
                tmp.append(self.boolean_instructor_subject[(i,s)])
            subject_diversity.append(sum(tmp))
        obj = self.model.NewIntVar(0,self.num_subject, "")
        self.model.AddMaxEquality(obj, subject_diversity)
        return obj
    
    # O-03
    def obj_quota_available(self):
        objective_quota_diff=[]
        for i in self.all_instructor:
            task_assigned=0
            for n in self.all_task:
                task_assigned+=self.assign[(n,i)]
            objective_quota_diff.append(self.instructor_quota[i]-task_assigned)
        obj = self.model.NewIntVar(0,self.num_task, "")
        self.model.AddMaxEquality(obj, objective_quota_diff)
        return obj
    
    # O-04
    def obj_walking_distance(self):
        walking_distance=[]
        for n1 in range(self.num_task-1):
            for n2 in range(n1+1,self.num_task):
                walking_distance.append(self.area_distance[self.task_area_mapping[n1]][self.task_area_mapping[n2]]*self.area_slot_weight[self.task_slot_mapping[n1]][self.task_slot_mapping[n2]]*self.boolean_product[(n1,n2)])
        return sum(walking_distance)
    
    # O-05
    def obj_subject_preference(self):
        objective_subject_preference=[]
        for i in self.all_instructor:
            for n in self.all_task:
                objective_subject_preference.append(self.instructor_subject_preference[i][self.task_subject_mapping[n]] * self.assign[(n, i)] )
        return sum( objective_subject_preference )
    
    # O-06
    def obj_slot_preference(self):
        objective_slot_preference=[]
        for i in self.all_instructor:
            for n in self.all_task:
                objective_slot_preference.append(self.instructor_slot_preference[i][self.task_slot_mapping[n]] * self.assign[(n, i)] )
        return sum( objective_slot_preference )
    '''
    ################################
    ||          Utility           ||
    ################################
    '''
    def create_delta(self,max_delta,actual_value,target_value):
        delta=self.model.NewIntVarFromDomain(cp_model.Domain.FromIntervals([[0, max_delta]]), '')
        self.model.Add(actual_value <= target_value + delta)
        self.model.Add(actual_value >= target_value - delta)
        return delta

    def optimizeObjectiveSolver(self):
        '''
        ################################
        ||   OBJECTIVE OPTIMIZATION   ||
        ################################
        '''
        # CREATE SOLVER
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = self.MAX_CP_TIME
        total_objective=[]     
        # O-03 OBJECTIVE MINIMIZE QUOTA DIFF
        if self.OBJ_STATUS[2]:
            total_objective.append(self.create_delta(self.num_task,self.obj_quota_available(),0 ))

        # O-05 OBJECTIVE MAXIMIZE SUBJECT PREFERENCE
        if self.OBJ_STATUS[4]:
            total_objective.append( self.create_delta(self.num_task*5,self.obj_subject_preference(),self.num_task*5) )
        
        # O-06 OBJECTIVE MAXIMIZE SLOT PREFERENCE
        if self.OBJ_STATUS[5]:
            total_objective.append( self.create_delta(self.num_task*5,self.obj_slot_preference(),self.num_task*5))
        
        # O-02 OBJECTIVE SUBJECT DIVERSITY
        if self.OBJ_STATUS[1]:
            self.boolean_instructor_subject={}
            for i in self.all_instructor:
                for s in self.all_subject:  
                    total=sum([self.assign[(n,i)] for n in self.all_task if self.task_subject_mapping[n] == s])                  
                    self.boolean_instructor_subject[(i,s)]=self.model.NewBoolVar('instructor_i%is%i' % (i,s))
                    # self.model.AddMinEquality(self.boolean_instructor_subject[(i,s)],[1,total])
                    self.model.Add( total > 0 ).OnlyEnforceIf( self.boolean_instructor_subject[(i,s)])
                    self.model.Add( total == 0 ).OnlyEnforceIf( self.boolean_instructor_subject[(i,s)].Not())
            total_objective.append(self.create_delta(self.num_subject,self.obj_subject_diversity(),0))
        # FOR O-01 O-04 Boolean Product Of assign (n1,i) and assign(n2,i)
        if self.OBJ_STATUS[0] or self.OBJ_STATUS[3]:
            self.boolean_product={}
            for n1 in range(self.num_task-1):
                for n2 in range(n1+1,self.num_task):
                    tmpvars=[]
                    self.boolean_product[(n1,n2)] = self.model.NewBoolVar('product_n1%in2%i' % (n1,n2))
                    for i in self.all_instructor:
                        tmpvar=self.model.NewBoolVar("")
                        self.model.AddMultiplicationEquality(tmpvar,[self.assign[(n1, i)],self.assign[(n2,i)]])
                        tmpvars.append(tmpvar)
                    self.model.AddMaxEquality(self.boolean_product[(n1,n2)],tmpvars)

        # O-01 OBJECTIVE MAXIMIZE TIMETABLE COMPATIBILITY
        if self.OBJ_STATUS[0]:
            total_objective.append ( self.create_delta(self.num_slot*(self.num_slot-1)*5*2,self.obj_slot_compatibility(),self.num_slot*(self.num_slot-1)*5) )
        # O-04 OBJECTIVE MINIMIZE WALKING DISTANCE
        if self.OBJ_STATUS[3]:
            total_objective.append( self.create_delta(self.num_task*5,self.obj_walking_distance(),0) )

        # CONSTRAINT PROGRAMMING
        self.model.Minimize(sum(total_objective))        
        status = solver.Solve(self.model)
        result = []
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            for n in self.all_task:
                for i in self.all_instructor:
                    if(solver.Value(self.assign[(n,i)])):
                        result.append(i)
                        break
        self.print_solution(status,solver,result)

    def constraintOnlySolver(self):
        '''
        ################################
        ||      CONSTRAINT ONLY       ||
        ################################
        '''
        solver = cp_model.CpSolver()
        solver.parameters.linearization_level = 0
        solver.parameters.max_time_in_seconds = self.MAX_CP_TIME
        self.model.Minimize(self.obj_backup_instructor())
        status = solver.Solve(self.model)
        result = []
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            for n in self.all_task:
                for i in self.all_instructor:
                    if(solver.Value(self.assign[(n,i)])):
                        result.append(i)
                        break
        self.print_solution(status,solver,result)

    def print_solution(self,status,solver,result):
        now = datetime.datetime.now()
        date_string = now.strftime("%Y-%m-%dT%H-%M-%S")   
        with open(f'results/results_{date_string}.txt',mode='w') as the_file:
            the_file.write(f"[Solver Information]\n")
            the_file.write(f'- Excel input path : {self.EXCEL_INPUT_PATH}\n')   
            the_file.write(f'- Max execute time : {self.MAX_CP_TIME}\n')   
            the_file.write(f'- Objective Status : {self.OBJ_STATUS}\n\n')
            the_file.write(f"[Solution]\n")
            the_file.write(f"<Statistics>\n")
            the_file.write(f'- Status : {solver.StatusName(status)}\n')
            the_file.write(f"- Objective : {solver.ObjectiveValue()}\n")   
            the_file.write(f'- Conflicts : {solver.NumConflicts()}\n')  
            the_file.write(f'- Branches : {solver.NumBranches()}\n')
            the_file.write(f'- Wall time : {solver.WallTime()} s\n')  
            the_file.write(f"<Result>\n")
            the_file.write(f"{result}\n")

    def solve(self):
        self.create_variable()
        self.create_model()
        if sum(OBJECTIVE_STATUS):
            self.optimizeObjectiveSolver()
        else:
            self.constraintOnlySolver()

if __name__ == '__main__':
    solver = IntructorAssigningSolver(max_time_per_obj=MAX_EXCECUTE_TIME,excel_input_path=EXCEL_INPUT_PATH,obj_status=OBJECTIVE_STATUS)
    solver.solve()