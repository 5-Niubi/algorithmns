from collections import Counter
import openpyxl

INPUT_EXCEL_PATH = "inputs/inputSE.xlsx"
RESULT=[14, 24, 17, 6, 18, 28, 15, 32, 26, 4, 22, 13, 14, 28, 17, 12, 13, 5, 22, 28, 0, 14, 24, 29, 22, 6, 26, 14, 14, 32, 29, 10, 17, 32, 7, 26, 22, 28, 13, 10, 7, 0, 5, 13, 3, 27, 32, 28, 10, 3, 0, 15, 12, 14, 27, 15, 11, 12, 8, 25, 13, 8, 10, 26, 8, 18, 5, 24, 17, 4, 18, 12, 0, 21, 23, 30, 13, 11, 30, 11, 11, 11, 30, 11, 16, 11, 2, 2, 21, 2, 2, 16, 21, 16, 11, 29, 30, 30, 6, 30, 21, 29, 6, 29, 6, 6, 30, 29, 4, 30, 4, 19, 13, 1, 23, 19, 21, 16, 16, 16, 21, 19, 1, 24, 33, 33, 11, 29, 1, 9, 25, 31, 25, 9, 8, 9, 23, 31, 3, 9, 15, 31, 21, 20, 31, 23, 1, 28, 28, 24, 23, 30, 7, 7, 19, 20, 5, 24, 27, 20, 27, 20]

'''
################################
||           SHEET            ||
################################
'''
data_wb = openpyxl.load_workbook(INPUT_EXCEL_PATH)
info_sheet = data_wb.worksheets[0]
task_sheet = data_wb.worksheets[1]
slot_conflict_sheet = data_wb.worksheets[2]
slot_compatibility_sheet = data_wb.worksheets[3]
instructor_ability_sheet = data_wb.worksheets[4]
instructor_slot_sheet = data_wb.worksheets[5]
instructor_quota_sheet = data_wb.worksheets[6]
instructor_preassign_sheet = data_wb.worksheets[7]
area_distance_sheet = data_wb.worksheets[8]
area_slot_weight_sheet = data_wb.worksheets[9]

num_task = info_sheet.cell(1,2).value
num_instructor = info_sheet.cell(2,2).value
num_slot = info_sheet.cell(3,2).value
num_subject = info_sheet.cell(4,2).value
num_area = info_sheet.cell(5,2).value

all_task = range(num_task)
all_instructor = range(num_instructor)
all_slot = range(num_slot)
all_subject = range(num_subject)
all_area = range(num_area)

slot_conflict = [[0] * num_slot for i in all_slot]
for i in all_slot:
    for j in all_slot:
        if slot_conflict_sheet.cell(i+2,j+2).value:
            slot_conflict[i][j]=1
slot_compatibility=[[0] * num_slot for i in all_slot]
for i in all_slot:
    for j in all_slot:
            slot_compatibility[i][j]=int(slot_compatibility_sheet.cell(i+2,j+2).value)

instructor_ability = [[0] * num_subject for i in all_instructor]
instructor_subject_preference = [[0] * num_subject for i in all_instructor]
for i in all_instructor:
    for j in all_subject:
        if instructor_ability_sheet.cell(i+2,j+2).value:
            instructor_ability[i][j]=1
        instructor_subject_preference[i][j]=instructor_ability_sheet.cell(i+2,j+2).value

instructor_slot = [[0] * num_slot for i in all_instructor]
instructor_slot_preference= [[0] * num_slot for i in all_instructor]
for i in all_instructor:
    for j in all_slot:
        if instructor_slot_sheet.cell(i+2,j+2).value:
            instructor_slot[i][j]=1
        instructor_slot_preference[i][j]=instructor_slot_sheet.cell(i+2,j+2).value

task_subject_mapping = []
for i in all_task:
    for j in all_subject:
        if task_sheet.cell(i+2,2).value == instructor_ability_sheet.cell(1,j+2).value:
            task_subject_mapping.append(j)

task_slot_mapping = []
for i in all_task:
    for j in all_slot:
        if task_sheet.cell(i+2,4).value == instructor_slot_sheet.cell(1,j+2).value:
            task_slot_mapping.append(j)

instructor_quota=[]
for i in all_instructor:
    instructor_quota.append(int ( instructor_quota_sheet.cell(i+2,2).value) )


instructor_preassign = [[0] * num_task for i in all_instructor]
for i in all_instructor:
    for n in all_task:
        if instructor_preassign_sheet.cell(i+2,n+2).value:
            instructor_preassign[i][instructor_preassign_sheet.cell(i+2,n+2).value-1]=1

area_distance = [[0]* num_area for i in all_area]
for i in all_area:
    for j in all_area:
        area_distance[i][j] = area_distance_sheet.cell(i+2,j+2).value

area_slot_weight = [[0] * num_slot for i in all_slot]
for i in all_slot:
    for j in all_slot:
        area_slot_weight[i][j]=int(area_slot_weight_sheet.cell(i+2,j+2).value)

task_area_mapping = []
for n in all_task:
    dept=task_sheet.cell(n+2,7).value[0:2]
    if(dept=="AL"): task_area_mapping.append(0)
    if(dept=="BE"): task_area_mapping.append(1)
    if(dept=="DE"): task_area_mapping.append(2)

assign=[[0]* (num_instructor+1) for n in all_task]

for idx,_ in enumerate(RESULT):
    assign[idx][_]=1

slot_task_listing = [[] for s in all_slot]
for n in all_task:
    slot_task_listing[task_slot_mapping[n]].append(n)


slot_task_conflicting = [[] for s in all_slot]
for j in all_slot:
    for n in all_task:
        if(slot_conflict[task_slot_mapping[n]][j]):
            slot_task_conflicting[j].append(n)

'''
################################
||         CONSTRAINT         ||
################################
'''
print("<Constraint>")
#C-00 Total Task Assigned
task_assigned = num_task - RESULT.count(num_instructor)
print(f"- Task Assigned : {task_assigned}/{num_task}")
#C-01 No Slot Conflict
slot_conflict_count=0
for i in all_instructor:
    for j in all_slot:     
        total_task_in_that_slot = 0
        for _ in slot_task_listing[j]:
            total_task_in_that_slot+=assign[_][i]

        if(total_task_in_that_slot > 0):
            total_task=0
            for _ in slot_task_conflicting[j]:
                total_task+=assign[_][i]
            if(total_task>2):
                slot_conflict_count+=1
print(f"- Slot Conflict : {slot_conflict_count}")
#C-02 Pre-assign must be satisfy
preassign_not_satisfy_count = 0
for i in all_instructor:
    for s in all_slot:
        n = instructor_preassign_sheet.cell(i+2,s+2).value
        if n is not None and RESULT[n]!=i:
            preassign_not_satisfy_count += 1
print(f"- Preassign Not Satisfy : {preassign_not_satisfy_count}")
#C-03 Instructor Subject
instructor_subject_error = 0
for n,i in enumerate(RESULT):
    if instructor_ability[i][task_subject_mapping[n]] == 0:
        instructor_subject_error += 1
print(f"- Instructor Subject Error : {instructor_subject_error}")
#C-04 Instructor Slot
instructor_slot_error = 0
for n,i in enumerate(RESULT):
    if instructor_slot[i][task_slot_mapping[n]] == 0:
        instructor_slot_error += 1
print(f"- Instructor Slot Error : {instructor_slot_error}")
'''
################################
||          Utility           ||
################################
'''
def groupByInstructor(solution):
    grouped={}
    for n in all_task:
        if solution[n] == num_instructor:
           continue
        if solution[n] not in grouped:
            grouped[solution[n]]=[n]
        else:
            grouped[solution[n]].append(n)
    return grouped
'''
################################
||         Objective          ||
################################
'''
print("<Objective>")
# O-01 Slot Compability Cost
def slotCompability(solution):
    total_cost=0
    grouped = groupByInstructor(solution)
    for tasks in grouped.values():
        for n1 in range(0,len(tasks)-1):
            for n2 in range(n1,len(tasks)):
                total_cost+=slot_compatibility[task_slot_mapping[tasks[n1]]][task_slot_mapping[tasks[n2]]]
    return total_cost

print(f"- Slot compability : {slotCompability(RESULT)}")
# O-02 Subject Diversity
def subjectDiversity(solution):
    grouped = groupByInstructor(solution)
    subjectDiversitys=[]
    for tasks in grouped.values():
        for idx,task in enumerate(tasks):
            tasks[idx]=task_subject_mapping[task]
        subjectDiversitys.append( len(set(tasks)))
    return max(subjectDiversitys)
print(f"- Subject Diversity : {subjectDiversity(RESULT)}")
# O-03 Quota Satisfy
def quotaAvailable(solution):
    total_error=0
    count = Counter(solution)
    for i in all_instructor:
        total_error =  max(total_error, instructor_quota[i] - count[i])
    return total_error
print(f"- Max Quota Available : {quotaAvailable(RESULT)}")
# 0-04 Walking Distance
def walkingDistance(solution):
    total_cost=0
    grouped = groupByInstructor(solution)
    for tasks in grouped.values():
        for n1 in range(0,len(tasks)-1):
            for n2 in range(n1,len(tasks)):
                total_cost+=area_slot_weight[task_slot_mapping[tasks[n1]]][task_slot_mapping[tasks[n2]]] * area_distance[task_area_mapping[tasks[n1]]][task_area_mapping[tasks[n2]]]
    return total_cost
print(f"- Walking Distance : {walkingDistance(RESULT)}")
# O-05 Subject Preference
def subjectPreference(solution):
    total_score=0
    for n in all_task:
        if(solution[n]!= num_instructor):
            total_score += instructor_subject_preference[solution[n]][task_subject_mapping[n]]
    return - total_score
print(f"- Subject Preference : {-subjectPreference(RESULT)}")
# O-06 Slot Preference
def slotPreference(solution):
    total_score=0
    for n in all_task:
        if(solution[n]!= num_instructor):
            total_score += instructor_slot_preference[solution[n]][task_slot_mapping[n]]
    return - total_score
print(f"- Slot Preference : {-slotPreference(RESULT)}")
