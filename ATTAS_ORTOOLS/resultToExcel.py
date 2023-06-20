import openpyxl
from openpyxl.styles import PatternFill,Border, Side,Alignment
import datetime

INPUT_EXCEL_PATH = "inputs/inputSE.xlsx"
RESULT = [14, 24, 17, 6, 18, 28, 15, 32, 26, 4, 22, 13, 14, 28, 17, 12, 13, 5, 22, 28, 0, 14, 24, 29, 22, 6, 26, 14, 14, 32, 29, 10, 17, 32, 7, 26, 22, 28, 13, 10, 7, 0, 5, 13, 3, 27, 32, 28, 10, 3, 0, 15, 12, 14, 27, 15, 11, 12, 8, 25, 13, 8, 10, 26, 8, 18, 5, 24, 17, 4, 18, 12, 0, 21, 23, 30, 13, 11, 30, 11, 11, 11, 30, 11, 16, 11, 2, 2, 21, 2, 2, 16, 21, 16, 11, 29, 30, 30, 6, 30, 21, 29, 6, 29, 6, 6, 30, 29, 4, 30, 4, 19, 13, 1, 23, 19, 21, 16, 16, 16, 21, 19, 1, 24, 33, 33, 11, 29, 1, 9, 25, 31, 25, 9, 8, 9, 23, 31, 3, 9, 15, 31, 21, 20, 31, 23, 1, 28, 28, 24, 23, 30, 7, 7, 19, 20, 5, 24, 27, 20, 27, 20]

'''
################################
||           SHEET            ||
################################
'''
data_wb = openpyxl.load_workbook(INPUT_EXCEL_PATH)
info_sheet = data_wb.worksheets[0]
task_sheet = data_wb.worksheets[1]
slot_conflict_sheet = data_wb.worksheets[2]
instructor_ability_sheet = data_wb.worksheets[4]
instructor_slot_sheet = data_wb.worksheets[5]

num_task = info_sheet.cell(1,2).value
num_instructor = info_sheet.cell(2,2).value
num_slot = info_sheet.cell(3,2).value
num_subject = info_sheet.cell(4,2).value

all_task = range(num_task)
all_slot = range(num_slot)

task_slot_mapping = []
for i in all_task:
    for j in all_slot:
        if task_sheet.cell(i+2,4).value == instructor_slot_sheet.cell(1,j+2).value:
            task_slot_mapping.append(j)

result_wb = openpyxl.Workbook()
result_sheet = result_wb.active

align_middle = Alignment(vertical='center',horizontal='center')
full_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
dark_orange = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
steel_blue = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
for i in range(1,num_instructor+3):
    for j in range(1,num_slot+2):
        result_sheet.cell(row=i,column=j).border = full_border

for i in range(2,num_instructor+2):
    result_sheet.cell(row=i,column=1,value=instructor_ability_sheet.cell(row=i,column=1).value)
    result_sheet.cell(row=i,column=1).fill=dark_orange
result_sheet.cell(row=num_instructor+2,column=1,value="UNASSIGNED")
result_sheet.cell(row=num_instructor+2,column=1).fill=dark_orange

for i in range(2,num_slot+2):
    result_sheet.cell(row=1,column=i,value=slot_conflict_sheet.cell(row=1,column=i).value)
    result_sheet.cell(row=1,column=i).fill=steel_blue
    result_sheet.cell(row=1,column=i).alignment=align_middle

for n in all_task:
    value = result_sheet.cell(row=RESULT[n]+2,column=task_slot_mapping[n]+2).value
    if value is not None:
        value+=f"\n{n+1}.{task_sheet.cell(row=n+2,column=1).value}.{task_sheet.cell(row=n+2,column=2).value}"
    else:
        value=f"{n+1}.{task_sheet.cell(row=n+2,column=1).value}.{task_sheet.cell(row=n+2,column=2).value}"
    result_sheet.cell(row=RESULT[n]+2,column=task_slot_mapping[n]+2,value=value)

for col in result_sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            # Get the length of the content of the cell
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Add some padding to the width
    result_sheet.column_dimensions[column].width = adjusted_width if adjusted_width > 5 else 5  # Set the width of the column, with a minimum of 5

now = datetime.datetime.now()
date_string = now.strftime("%Y-%m-%dT%H-%M-%S")
result_wb.save(f"results/result_{date_string}.xlsx")